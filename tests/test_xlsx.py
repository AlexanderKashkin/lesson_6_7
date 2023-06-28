import os
from openpyxl import load_workbook
from path import dir_main


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx():
    path_for_xlsx = os.path.join(dir_main, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(path_for_xlsx)
    sheet = workbook.active
    data = sheet.cell(row=3, column=2).value
    assert data == 'Mara', 'data is not Mara'
    assert os.path.getsize(path_for_xlsx) == 7360, 'size is not 7360'
